from openpyxl import Workbook

from utils.excel_reader import read_test_cases
from utils.excel_writer import update_test_result


def _make_workbook(tmp_path, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Test ID", "Question", "Expected Answer", "Actual Answer", "Result"])
    for row in rows:
        sheet.append(row)

    file_path = tmp_path / "TestData.xlsx"
    workbook.save(file_path)
    return file_path


def test_read_test_cases_returns_one_entry_per_row(tmp_path):
    # Arrange
    file_path = _make_workbook(
        tmp_path,
        rows=[
            ["TC001", "What is the capital of Saudi Arabia?", "Riyadh", "", ""],
            ["TC002", "What is 10 + 20?", "30", "", ""],
        ],
    )

    # Act
    test_cases = read_test_cases(file_path)

    # Assert
    assert len(test_cases) == 2
    assert test_cases[0].test_id == "TC001"
    assert test_cases[0].question == "What is the capital of Saudi Arabia?"
    assert test_cases[0].expected_answer == "Riyadh"


def test_read_test_cases_skips_blank_rows(tmp_path):
    # Arrange
    file_path = _make_workbook(
        tmp_path,
        rows=[
            ["TC001", "What is the capital of Saudi Arabia?", "Riyadh", "", ""],
            [None, None, None, None, None],
            ["TC002", "What is 10 + 20?", "30", "", ""],
        ],
    )

    # Act
    test_cases = read_test_cases(file_path)

    # Assert
    assert len(test_cases) == 2
    assert [tc.test_id for tc in test_cases] == ["TC001", "TC002"]


def test_update_test_result_writes_actual_answer_and_result(tmp_path):
    # Arrange
    file_path = _make_workbook(
        tmp_path,
        rows=[["TC001", "What is the capital of Saudi Arabia?", "Riyadh", "", ""]],
    )

    # Act
    update_test_result(file_path, row_number=2, actual_answer="Riyadh", result="PASS")

    # Assert
    updated_cases = read_test_cases(file_path)
    assert updated_cases[0].test_id == "TC001"

    from openpyxl import load_workbook
    workbook = load_workbook(file_path)
    sheet = workbook.active
    assert sheet.cell(row=2, column=4).value == "Riyadh"
    assert sheet.cell(row=2, column=5).value == "PASS"
