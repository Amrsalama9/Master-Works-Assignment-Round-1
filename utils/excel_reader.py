from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from config.settings import (
    COLUMN_TEST_ID,
    COLUMN_QUESTION,
    COLUMN_EXPECTED_ANSWER,
)


@dataclass
class TestCase:
    row_number: int  # 1-based Excel row, needed later so the writer knows where to update
    test_id: str
    question: str
    expected_answer: str


def get_column_index(sheet) -> dict[str, int]:
    """
    Maps header name -> 1-based column index. Shared with excel_writer so
    both modules agree on where each column lives without duplicating the
    lookup logic.
    """
    header_row = [cell.value for cell in sheet[1]]
    return {name: header_row.index(name) + 1 for name in header_row if name}


def read_test_cases(excel_path: Path) -> list[TestCase]:
    """
    Reads every row from the first sheet and returns one TestCase per row.
    Assumes row 1 is the header row, matching the layout given in the
    assignment (Test ID, Question, Expected Answer, Actual Answer, Result).

    Rows with an empty Test ID are skipped - this lets someone leave a
    blank row in the sheet without it turning into a bogus test case.
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"Test data file not found: {excel_path}")

    workbook = load_workbook(excel_path)
    sheet = workbook.active

    column_index = get_column_index(sheet)

    for required_column in (COLUMN_TEST_ID, COLUMN_QUESTION, COLUMN_EXPECTED_ANSWER):
        if required_column not in column_index:
            raise ValueError(f"'{required_column}' column not found in {excel_path.name}")

    test_cases = []
    for row_number in range(2, sheet.max_row + 1):
        test_id = sheet.cell(row=row_number, column=column_index[COLUMN_TEST_ID]).value
        if not test_id:
            continue

        question = sheet.cell(row=row_number, column=column_index[COLUMN_QUESTION]).value
        expected_answer = sheet.cell(
            row=row_number, column=column_index[COLUMN_EXPECTED_ANSWER]
        ).value

        test_cases.append(
            TestCase(
                row_number=row_number,
                test_id=str(test_id).strip(),
                question=str(question).strip() if question else "",
                expected_answer=str(expected_answer).strip() if expected_answer else "",
            )
        )

    workbook.close()
    return test_cases
