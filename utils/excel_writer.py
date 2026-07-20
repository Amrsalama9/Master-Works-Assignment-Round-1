from pathlib import Path

from openpyxl import load_workbook

from config.settings import COLUMN_ACTUAL_ANSWER, COLUMN_RESULT
from utils.excel_reader import get_column_index


def update_test_result(excel_path: Path, row_number: int, actual_answer: str, result: str) -> None:
    """
    Writes the actual answer and PASS/FAIL result for one row and saves
    the workbook immediately. Saving after every row (instead of batching
    at the end) means a crash mid-run doesn't lose results already
    collected.
    """
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    column_index = get_column_index(sheet)

    sheet.cell(row=row_number, column=column_index[COLUMN_ACTUAL_ANSWER]).value = actual_answer
    sheet.cell(row=row_number, column=column_index[COLUMN_RESULT]).value = result

    workbook.save(excel_path)
    workbook.close()
